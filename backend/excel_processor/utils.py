import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any, Tuple
import os
from io import BytesIO
import hashlib


def extract_headers_from_excel(file_path: str) -> Tuple[List[str], pd.DataFrame]:
    """
    Extract headers and data from Excel file
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        tuple of (headers list, DataFrame)
    """
    try:
        # Try reading as Excel
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path, engine='openpyxl')
        
        # Get headers
        headers = df.columns.tolist()
        
        # Convert column names to strings
        headers = [str(h) for h in headers]
        
        return headers, df
        
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {str(e)}")


def infer_data_types(df: pd.DataFrame) -> Dict[str, str]:
    """
    Infer data types for each column
    
    Args:
        df: pandas DataFrame
        
    Returns:
        dict mapping column names to inferred types
    """
    type_mapping = {}
    
    for col in df.columns:
        # Get non-null values
        non_null = df[col].dropna()
        
        if len(non_null) == 0:
            type_mapping[col] = 'text'
            continue
        
        # Check if numeric
        if pd.api.types.is_numeric_dtype(df[col]):
            if pd.api.types.is_integer_dtype(df[col]):
                type_mapping[col] = 'integer'
            else:
                type_mapping[col] = 'decimal'
        # Check if datetime
        elif pd.api.types.is_datetime64_any_dtype(df[col]):
            type_mapping[col] = 'date'
        else:
            type_mapping[col] = 'text'
    
    return type_mapping


def hash_field_value(value: Any) -> str:
    """
    Hash a field value using SHA256
    
    Args:
        value: The value to hash
        
    Returns:
        SHA256 hash of the value as hex string
    """
    if pd.isna(value) or value is None or value == '':
        return ''  # Don't hash empty values
    
    # Convert value to string and encode to bytes
    value_str = str(value)
    value_bytes = value_str.encode('utf-8')
    
    # Generate SHA256 hash
    hash_object = hashlib.sha256(value_bytes)
    return hash_object.hexdigest()


def generate_validated_excel(
    original_file_path: str,
    mapped_headers: Dict[str, str],
    validation_report: Dict[str, Any],
    output_path: str
) -> str:
    """
    Generate new Excel file with mapped headers and validation errors as comma-separated values
    
    Args:
        original_file_path: Path to original Excel file
        mapped_headers: dict mapping original to new headers
        validation_report: validation results from validators
        output_path: path where to save output file
        
    Returns:
        path to generated file
    """
    # Read original data
    if original_file_path.endswith('.csv'):
        df = pd.read_csv(original_file_path)
    else:
        df = pd.read_excel(original_file_path, engine='openpyxl')
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Create data sheet with mapped headers
    ws_data = wb.create_sheet("Verified Data")
    
    # Rename columns
    df_renamed = df.rename(columns=mapped_headers)
    
    # Create error mapping: row_number -> list of error messages
    error_messages = {}
    for error in validation_report.get('errors', []):
        if 'row' in error:
            row_num = error['row']
            error_msg = f"{error.get('column', 'Unknown')}: {error.get('issue', 'Error')}"
            if row_num not in error_messages:
                error_messages[row_num] = []
            error_messages[row_num].append(error_msg)
        elif 'rows' in error:
            for row_num in error.get('rows', []):
                error_msg = f"{error.get('column', 'Unknown')}: {error.get('issue', 'Error')}"
                if row_num not in error_messages:
                    error_messages[row_num] = []
                error_messages[row_num].append(error_msg)
    
    # Add "Validation Errors" column to dataframe
    df_renamed['Validation Errors'] = ''
    for row_num, messages in error_messages.items():
        # row_num is Excel row (1-indexed with header), dataframe index is 0-indexed
        df_index = row_num - 2  # Subtract 2 (1 for Excel 1-index, 1 for header row)
        if 0 <= df_index < len(df_renamed):
            df_renamed.loc[df_index, 'Validation Errors'] = ', '.join(messages)
    
    # Write data to sheet
    for r_idx, row in enumerate(dataframe_to_rows(df_renamed, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
            
            # Style header row
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Color-code error rows (red)
    error_rows = set()
    for error in validation_report.get('errors', []):
        if 'row' in error:
            error_rows.add(error['row'])
        elif 'rows' in error:
            error_rows.update(error['rows'])
    
    # Apply row colors (Excel row index = data row + 1 for header)
    for row_num in error_rows:
        for cell in ws_data[row_num]:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)  # Increased width for errors column
        ws_data.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook (no separate validation report sheet)
    wb.save(output_path)
    
    return output_path
    
    # Write summary
    summary_data = [
        ["Validation Summary", ""],
        ["Total Rows", len(df)],
        ["Valid Rows", validation_report.get('valid_rows_count', 0)],
        ["Invalid Rows", validation_report.get('invalid_rows_count', 0)],
        ["Errors", len(validation_report.get('errors', []))],
        ["Warnings", len(validation_report.get('warnings', []))],
        ["", ""],
        ["Errors Details", ""],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_report.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1 or row_idx == 8:
                cell.font = Font(bold=True, size=12)
    
    # Write errors
    current_row = len(summary_data) + 1
    if validation_report.get('errors'):
        ws_report.cell(row=current_row, column=1, value="Row").font = Font(bold=True)
        ws_report.cell(row=current_row, column=2, value="Column").font = Font(bold=True)
        ws_report.cell(row=current_row, column=3, value="Issue").font = Font(bold=True)
        ws_report.cell(row=current_row, column=4, value="Value").font = Font(bold=True)
        current_row += 1
        
        for error in validation_report.get('errors', []):
            row_info = str(error.get('row', '') or ', '.join(map(str, error.get('rows', []))))
            ws_report.cell(row=current_row, column=1, value=row_info)
            ws_report.cell(row=current_row, column=2, value=error.get('column', ''))
            ws_report.cell(row=current_row, column=3, value=error.get('issue', ''))
            ws_report.cell(row=current_row, column=4, value=error.get('value', ''))
            current_row += 1
    
    # Write warnings
    current_row += 1
    ws_report.cell(row=current_row, column=1, value="Warnings Details").font = Font(bold=True, size=12)
    current_row += 1
    
    if validation_report.get('warnings'):
        ws_report.cell(row=current_row, column=1, value="Column").font = Font(bold=True)
        ws_report.cell(row=current_row, column=2, value="Issue").font = Font(bold=True)
        ws_report.cell(row=current_row, column=3, value="Affected Rows").font = Font(bold=True)
        current_row += 1
        
        for warning in validation_report.get('warnings', []):
            ws_report.cell(row=current_row, column=1, value=warning.get('column', ''))
            ws_report.cell(row=current_row, column=2, value=warning.get('issue', ''))
            rows_str = ', '.join(map(str, warning.get('rows', [])[:10]))  # First 10 rows
            if len(warning.get('rows', [])) > 10:
                rows_str += '...'
            ws_report.cell(row=current_row, column=3, value=rows_str)
            current_row += 1
    
    # Auto-adjust report column widths
    for column in ws_report.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        ws_report.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_path)
    
    return output_path


def generate_good_and_rejected_files(
    original_file_path: str,
    mapped_headers: Dict[str, str],
    validation_report: Dict[str, Any],
    good_records_path: str,
    rejected_records_path: str,
    validation_options: Dict[str, bool] = None
) -> Tuple[str, str]:
    """
    Generate two separate Excel files: one with good records, one with rejected records
    
    Args:
        original_file_path: Path to original Excel file
        mapped_headers: dict mapping original to new headers
        validation_report: validation results from validators
        good_records_path: path where to save good records file
        rejected_records_path: path where to save rejected records file
        validation_options: dict with options like protect_good_records
        
    Returns:
        tuple of (good_records_path, rejected_records_path)
    """
    if validation_options is None:
        validation_options = {}
    
    # Read original data
    if original_file_path.endswith('.csv'):
        df = pd.read_csv(original_file_path)
    else:
        df = pd.read_excel(original_file_path, engine='openpyxl')
    
    # Rename columns
    df_renamed = df.rename(columns=mapped_headers)
    
    # Create error mapping: row_number -> list of error messages
    error_rows = set()
    error_messages = {}
    
    for error in validation_report.get('errors', []):
        if 'row' in error:
            row_num = error['row']
            error_rows.add(row_num)
            error_msg = f"{error.get('column', 'Unknown')}: {error.get('issue', 'Error')}"
            if row_num not in error_messages:
                error_messages[row_num] = []
            error_messages[row_num].append(error_msg)
        elif 'rows' in error:
            for row_num in error.get('rows', []):
                error_rows.add(row_num)
                error_msg = f"{error.get('column', 'Unknown')}: {error.get('issue', 'Error')}"
                if row_num not in error_messages:
                    error_messages[row_num] = []
                error_messages[row_num].append(error_msg)
    
    # Separate good and rejected records
    # error_rows are 1-indexed Excel rows (row 2 = first data row)
    # Convert to 0-indexed dataframe indices
    rejected_indices_set = {row_num - 2 for row_num in error_rows if 0 <= row_num - 2 < len(df_renamed)}
    
    # Use boolean indexing for better performance with large datasets
    is_rejected = [i in rejected_indices_set for i in range(len(df_renamed))]
    is_good = [not rejected for rejected in is_rejected]
    
    df_good = df_renamed[is_good].reset_index(drop=True)
    df_rejected = df_renamed[is_rejected].reset_index(drop=True)
    
    # Apply SHA256 hashing to specified fields in good records
    hash_fields = validation_options.get('hash_fields', []) if validation_options else []
    if hash_fields and len(df_good) > 0:
        for field in hash_fields:
            if field in df_good.columns:
                df_good[field] = df_good[field].apply(hash_field_value)
    
    # Keep rejected_indices as list for error message mapping
    rejected_indices = [i for i, rejected in enumerate(is_rejected) if rejected]
    
    # Add error column to rejected records
    df_rejected['Validation Errors'] = ''
    for idx, original_idx in enumerate(rejected_indices):
        row_num = original_idx + 2  # Convert back to Excel row number
        if row_num in error_messages:
            df_rejected.loc[idx, 'Validation Errors'] = ', '.join(error_messages[row_num])
    
    # Generate Good Records Excel
    if len(df_good) > 0:
        wb_good = openpyxl.Workbook()
        if 'Sheet' in wb_good.sheetnames:
            del wb_good['Sheet']
        ws_good = wb_good.create_sheet("Good Records")
        
        for r_idx, row in enumerate(dataframe_to_rows(df_good, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_good.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws_good.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_good.column_dimensions[column_letter].width = adjusted_width
        
        # Apply worksheet protection if option is enabled
        if validation_options and validation_options.get('protect_good_records', False):
            ws_good.protection.sheet = True
            ws_good.protection.enable()
            # Lock all cells
            for row in ws_good.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=True)
        
        wb_good.save(good_records_path)
    else:
        # Create empty file with headers only
        wb_good = openpyxl.Workbook()
        if 'Sheet' in wb_good.sheetnames:
            del wb_good['Sheet']
        ws_good = wb_good.create_sheet("Good Records")
        for c_idx, header in enumerate(df_renamed.columns, 1):
            cell = ws_good.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        
        # Apply protection even for empty file
        if validation_options and validation_options.get('protect_good_records', False):
            ws_good.protection.sheet = True
            ws_good.protection.enable()
        
        wb_good.save(good_records_path)
    
    # Generate Rejected Records Excel
    if len(df_rejected) > 0:
        wb_rejected = openpyxl.Workbook()
        if 'Sheet' in wb_rejected.sheetnames:
            del wb_rejected['Sheet']
        ws_rejected = wb_rejected.create_sheet("Rejected Records")
        
        for r_idx, row in enumerate(dataframe_to_rows(df_rejected, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_rejected.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # Highlight data rows with light red
                    cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws_rejected.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_rejected.column_dimensions[column_letter].width = adjusted_width
        
        wb_rejected.save(rejected_records_path)
    else:
        # Create empty file with headers
        wb_rejected = openpyxl.Workbook()
        if 'Sheet' in wb_rejected.sheetnames:
            del wb_rejected['Sheet']
        ws_rejected = wb_rejected.create_sheet("Rejected Records")
        for c_idx, header in enumerate(list(df_renamed.columns) + ['Validation Errors'], 1):
            cell = ws_rejected.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        wb_rejected.save(rejected_records_path)
    
    return good_records_path, rejected_records_path


def get_preview_data(df: pd.DataFrame, mapped_headers: Dict[str, str], num_rows: int = 50) -> List[Dict]:
    """
    Get preview data with mapped headers
    
    Args:
        df: pandas DataFrame
        mapped_headers: dict mapping original to new headers
        num_rows: number of rows to preview
        
    Returns:
        list of dicts representing rows
    """
    # Rename columns
    df_preview = df.head(num_rows).rename(columns=mapped_headers)
    
    # Convert to list of dicts
    preview_data = df_preview.to_dict('records')
    
    # Convert any NaN values to None
    for row in preview_data:
        for key in row:
            if pd.isna(row[key]):
                row[key] = None
    
    return preview_data
